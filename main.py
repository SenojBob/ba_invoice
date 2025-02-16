import openpyxl 
import os 

def main():
    excel_path = input("Enter the path of the Excel file: ").replace('"', '').strip()
    if os.path.exists(excel_path):
        workbook_r = openpyxl.load_workbook(excel_path, data_only=True)
    else:
        print(f"Workbook '{excel_path}' does not exist.")
    
    sheet = select_sheet(workbook_r)

    lookups_sheet = workbook_r['LookUps']
    suppliers = get_suppliers(lookups_sheet)
    path_name = f"{sheet['E8'].value[:3]} {sheet['E7'].value} - SUPPLIER - BA Invoice Breakdown.xlsx"
    folder = f"{sheet['E8'].value[:3]} {sheet['E7'].value}"
    path_name = os.path.join(folder, path_name)

    # Create folder for invoices
    if not os.path.exists(folder):
        os.makedirs(folder)

    row_lookup = {}
    for supplier in suppliers:
        row_lookup[supplier] = []

    row = 13
    val = sheet[f'E{row}'].value
    while val:
        supplier = sheet[f'E{row}'].value
        row_lookup[supplier].append(row)
        row += 1
        val = sheet[f'E{row}'].value

    for supplier in suppliers:
        print(f"Creating invoice for {supplier}")
        invoice = create_invoice(sheet, row_lookup[supplier], supplier)
        invoice_path = path_name.replace("SUPPLIER", supplier)
        invoice.save(invoice_path)
        invoice.close()

    print("Invoices created successfully.")

def select_sheet(workbook):
    print("Select sheet to create report from:")
    for idx, sheet in enumerate(workbook.sheetnames):
        if idx % 2 == 0:
            print(f"{idx + 1}: {sheet}".ljust(35), end='')
        else:
            print(f"{idx + 1}: {sheet}")
    
    sheet_index = input("Select the sheet number [Default: Invoice Template]: ")
    if not sheet_index:
        print("Selected default sheet: Invoice Template")
        return workbook['Invoice Template']
    sheet_index = int(sheet_index) - 1
    if 0 <= sheet_index < len(workbook.sheetnames):
        selected_sheet = workbook.sheetnames[sheet_index]
        print(f"Selected sheet: {selected_sheet}")
        return workbook[selected_sheet]
    else:
        print("Invalid sheet number selected.")
        exit()

def get_suppliers(lookups_sheet) -> list:
    suppliers = []
    for i in range(3, 50):
        supplier = lookups_sheet[f'K{i}'].value
        if supplier:
            suppliers.append(supplier)
    return suppliers

def create_invoice(vendor_sheet, rows, supplier) -> openpyxl.Workbook:
    # Open the template workbook
    template = openpyxl.load_workbook("Invoice_Template.xlsx")
    sheet = template.active

    # Copy the data from the vendor report to the invoice template
    # Header section
    for i in range(2, 5):
        sheet[f"B{i}"].value = vendor_sheet[f"E{i}"].value
    sheet["B5"].value = supplier
    sheet["B1"].value = vendor_sheet["E8"].value
    sheet["A7"].value = vendor_sheet["D5"].value
    sheet["B9"].value = vendor_sheet["E7"].value
    sheet["B10"].value = vendor_sheet["E8"].value

    # Line items
    row = 15
    for r in rows:
        sheet.insert_rows(row)
        curr_col = 1
        for i in range(1, 21):
            if i == 10:
                continue
            sheet.cell(row, curr_col).value = vendor_sheet.cell(r, i + 3).value
            format_cell(sheet.cell(row, curr_col))
            curr_col += 1
        row += 1

    row += 1

    # Total section
    for col in range(1, 20):
        cell = sheet.cell(row, col)
        cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    sheet.merge_cells(f"A{row}:K{row}")
    cell = sheet[f"A{row}"]
    cell.value = "SUBTOTAL"
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    for char in ['L', 'P', 'Q', 'R', 'S']:
        cell = sheet[f"{char}{row}"]
        cell.value = f"=+SUBTOTAL(9,{char}15:{char}{row - 1})"
        format_cell(cell)

    # Save the new workbook
    return template

def format_cell(cell):
    date_cols = [8, 9]
    acc_cols = [12, 16, 17, 18, 19]
    perc_cols = [13, 14, 15]

    col = cell.column
    if col in date_cols:
        cell.number_format = "mm/dd/yyyy"
    elif col in acc_cols:
        cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    elif col in perc_cols:
        cell.number_format = "0.00%"
    else:
        pass

if __name__ == '__main__':
    main()